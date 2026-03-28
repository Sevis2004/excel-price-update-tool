import os
import sys
from datetime import datetime

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

base_dir = os.path.dirname(__file__)
excel_file = os.path.join(base_dir, "input.xlsx")

workbook = load_workbook(excel_file)
sheet = workbook.active

for row in range(2, sheet.max_row + 1):
    url = sheet[f"A{row}"].value

    if not url:
        sheet[f"D{row}"] = "empty url"
        continue

    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        response.encoding = "utf-8"

        soup = BeautifulSoup(response.text, "html.parser")
        price_element = soup.select_one(".price_color")

        if price_element:
            price_text = price_element.get_text(strip=True)
            sheet[f"B{row}"] = price_text
            sheet[f"C{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sheet[f"D{row}"] = "success"
            print(f"Row {row}: success -> {price_text}")
        else:
            sheet[f"C{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sheet[f"D{row}"] = "price not found"
            print(f"Row {row}: price not found")

    except requests.RequestException:
        sheet[f"C{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet[f"D{row}"] = "request failed"
        print(f"Row {row}: request failed")

workbook.save(excel_file)
print("Done.")